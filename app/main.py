import csv
import io
import os
import secrets
from datetime import date, datetime
from pathlib import Path

from fastapi import FastAPI, Depends, HTTPException, UploadFile, File, Form, Request
from fastapi.responses import StreamingResponse, HTMLResponse, RedirectResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from starlette.middleware.sessions import SessionMiddleware
from sqlalchemy.orm import Session

from .db import Base, engine, get_db
from . import models, schemas, auth
from .engine import calculate_payroll, DEFAULT_RULE_CONFIG

Base.metadata.create_all(bind=engine)

app = FastAPI(title="Attendance Payroll Calculator")
app.add_middleware(
    SessionMiddleware,
    secret_key=os.environ.get("SESSION_SECRET", secrets.token_hex(32)),
    same_site="lax",
)

STATIC_DIR = Path(__file__).parent / "static"
app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")


@app.on_event("startup")
def bootstrap():
    from .seed import seed
    seed()


# ---------------------------------------------------------------------------
# Helpers — enforce per-tenant scoping
# ---------------------------------------------------------------------------
def _owned_location(db: Session, user: models.User, loc_id: int) -> models.Location:
    loc = db.get(models.Location, loc_id)
    if not loc or loc.company_id != user.company_id:
        raise HTTPException(404, "Location not found")
    return loc


def _owned_employee(db: Session, user: models.User, emp_id: int) -> models.Employee:
    emp = db.get(models.Employee, emp_id)
    if not emp:
        raise HTTPException(404)
    _owned_location(db, user, emp.location_id)
    return emp


# ---------------------------------------------------------------------------
# Auth
# ---------------------------------------------------------------------------
@app.post("/api/auth/login")
async def login(request: Request, db: Session = Depends(get_db)):
    body = await request.json()
    email = (body.get("email") or body.get("username") or "").strip().lower()
    user = db.query(models.User).filter_by(email=email).first()
    if not user or not auth.verify_password(body.get("password", ""), user.password_hash):
        raise HTTPException(401, "Invalid credentials")
    request.session["uid"] = user.id
    return {"ok": True, "email": user.email, "company_id": user.company_id}


@app.post("/api/auth/logout")
def logout(request: Request):
    request.session.clear()
    return {"ok": True}


@app.get("/api/auth/me")
def me(user: models.User = Depends(auth.get_current_user)):
    return {"id": user.id, "email": user.email, "company_id": user.company_id,
            "role": user.role, "is_admin": (user.role or "").lower() == "admin"}


_EMAIL_RE = __import__("re").compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


def _validate_email(s: str) -> str:
    s = (s or "").strip().lower()
    if not _EMAIL_RE.match(s):
        raise HTTPException(400, "A valid email address is required")
    return s


@app.post("/api/auth/register")
async def register(request: Request, db: Session = Depends(get_db)):
    """Create a new company + its first admin user in one step."""
    body = await request.json()
    company_name = body.get("company_name", "").strip()
    email = _validate_email(body.get("email") or body.get("username"))
    password = body.get("password", "")
    if not (company_name and password):
        raise HTTPException(400, "company_name, email, password required")
    if db.query(models.Company).filter_by(name=company_name).first():
        raise HTTPException(400, "Company already exists")
    if db.query(models.User).filter_by(email=email).first():
        raise HTTPException(400, "Email already registered")
    comp = models.Company(name=company_name)
    db.add(comp); db.flush()
    user = models.User(company_id=comp.id, email=email,
                       password_hash=auth.hash_password(password),
                       role="admin", is_admin=True)
    db.add(user); db.commit()
    request.session["uid"] = user.id
    return {"ok": True, "company_id": comp.id, "email": email, "role": "admin"}


# ---------- User management (admin-only) ----------
@app.get("/api/users", response_model=list[schemas.UserOut])
def list_users(user: models.User = Depends(auth.require_admin),
               db: Session = Depends(get_db)):
    return db.query(models.User).filter_by(company_id=user.company_id).all()


@app.post("/api/users", response_model=schemas.UserOut)
def create_user(body: schemas.UserCreate,
                user: models.User = Depends(auth.require_admin),
                db: Session = Depends(get_db)):
    email = _validate_email(body.email)
    if db.query(models.User).filter_by(email=email).first():
        raise HTTPException(400, "Email already registered")
    role = (body.role or "kam").lower()
    if role not in ("admin", "kam"):
        raise HTTPException(400, "Role must be admin or kam")
    u = models.User(
        company_id=user.company_id, email=email,
        password_hash=auth.hash_password(body.password),
        role=role, is_admin=(role == "admin"),
    )
    db.add(u); db.commit(); db.refresh(u)
    return u


@app.delete("/api/users/{uid}")
def delete_user(uid: int,
                user: models.User = Depends(auth.require_admin),
                db: Session = Depends(get_db)):
    target = db.get(models.User, uid)
    if not target or target.company_id != user.company_id:
        raise HTTPException(404)
    if target.id == user.id:
        raise HTTPException(400, "Cannot delete self")
    db.delete(target); db.commit()
    return {"ok": True}


# ---------------------------------------------------------------------------
# Root
# ---------------------------------------------------------------------------
@app.get("/", response_class=HTMLResponse)
def index():
    return (STATIC_DIR / "index.html").read_text()


@app.get("/login", response_class=HTMLResponse)
def login_page():
    return (STATIC_DIR / "login.html").read_text()


@app.get("/api/default-config")
def default_config(user: models.User = Depends(auth.get_current_user)):
    return DEFAULT_RULE_CONFIG


# ---------------------------------------------------------------------------
# Locations
# ---------------------------------------------------------------------------
@app.post("/api/locations", response_model=schemas.LocationOut)
def create_location(body: schemas.LocationIn,
                    user: models.User = Depends(auth.require_admin),
                    db: Session = Depends(get_db)):
    data = body.model_dump()
    data["company_id"] = user.company_id  # force scope
    loc = models.Location(**data)
    db.add(loc); db.commit(); db.refresh(loc)
    return loc


@app.put("/api/locations/{loc_id}", response_model=schemas.LocationOut)
def update_location(loc_id: int, body: schemas.LocationIn,
                    user: models.User = Depends(auth.require_admin),
                    db: Session = Depends(get_db)):
    loc = _owned_location(db, user, loc_id)
    for k, v in body.model_dump().items():
        if k == "company_id":
            continue  # never reassign
        setattr(loc, k, v)
    db.commit(); db.refresh(loc)
    return loc


@app.delete("/api/locations/{loc_id}")
def delete_location(loc_id: int,
                    user: models.User = Depends(auth.require_admin),
                    db: Session = Depends(get_db)):
    loc = _owned_location(db, user, loc_id)
    db.delete(loc); db.commit()
    return {"ok": True}


@app.get("/api/locations", response_model=list[schemas.LocationOut])
def list_locations(user: models.User = Depends(auth.get_current_user),
                   db: Session = Depends(get_db)):
    return db.query(models.Location).filter_by(company_id=user.company_id).all()


@app.get("/api/locations/{loc_id}", response_model=schemas.LocationOut)
def get_location(loc_id: int,
                 user: models.User = Depends(auth.get_current_user),
                 db: Session = Depends(get_db)):
    return _owned_location(db, user, loc_id)


# ---------------------------------------------------------------------------
# Employees
# ---------------------------------------------------------------------------
@app.post("/api/employees", response_model=schemas.EmployeeOut)
def create_employee(body: schemas.EmployeeIn,
                    user: models.User = Depends(auth.get_current_user),
                    db: Session = Depends(get_db)):
    _owned_location(db, user, body.location_id)
    e = models.Employee(**body.model_dump())
    db.add(e); db.commit(); db.refresh(e)
    return e


@app.put("/api/employees/{emp_id}", response_model=schemas.EmployeeOut)
def update_employee(emp_id: int, body: schemas.EmployeeIn,
                    user: models.User = Depends(auth.get_current_user),
                    db: Session = Depends(get_db)):
    emp = _owned_employee(db, user, emp_id)
    _owned_location(db, user, body.location_id)
    for k, v in body.model_dump().items():
        setattr(emp, k, v)
    db.commit(); db.refresh(emp)
    return emp


@app.delete("/api/employees/{emp_id}")
def delete_employee(emp_id: int,
                    user: models.User = Depends(auth.get_current_user),
                    db: Session = Depends(get_db)):
    emp = _owned_employee(db, user, emp_id)
    db.delete(emp); db.commit()
    return {"ok": True}


@app.get("/api/employees", response_model=list[schemas.EmployeeOut])
def list_employees(location_id: int | None = None,
                   user: models.User = Depends(auth.get_current_user),
                   db: Session = Depends(get_db)):
    loc_ids = [l.id for l in db.query(models.Location)
               .filter_by(company_id=user.company_id).all()]
    q = db.query(models.Employee).filter(models.Employee.location_id.in_(loc_ids))
    if location_id:
        _owned_location(db, user, location_id)
        q = q.filter(models.Employee.location_id == location_id)
    return q.all()


# ---------------------------------------------------------------------------
# Holiday calendar
# ---------------------------------------------------------------------------
@app.get("/api/holidays")
def list_holidays(location_id: int,
                  user: models.User = Depends(auth.get_current_user),
                  db: Session = Depends(get_db)):
    _owned_location(db, user, location_id)
    hs = db.query(models.Holiday).filter_by(location_id=location_id).order_by(models.Holiday.day).all()
    return [{"id": h.id, "day": h.day.isoformat(), "name": h.name} for h in hs]


@app.post("/api/holidays")
async def add_holiday(request: Request,
                      user: models.User = Depends(auth.get_current_user),
                      db: Session = Depends(get_db)):
    body = await request.json()
    location_id = int(body["location_id"])
    _owned_location(db, user, location_id)
    d = datetime.strptime(body["day"], "%Y-%m-%d").date()
    existing = db.query(models.Holiday).filter_by(location_id=location_id, day=d).first()
    if existing:
        existing.name = body.get("name", existing.name)
    else:
        db.add(models.Holiday(location_id=location_id, day=d, name=body.get("name", "Holiday")))
    db.commit()
    return {"ok": True}


@app.delete("/api/holidays/{hid}")
def delete_holiday(hid: int,
                   user: models.User = Depends(auth.get_current_user),
                   db: Session = Depends(get_db)):
    h = db.get(models.Holiday, hid)
    if not h:
        raise HTTPException(404)
    _owned_location(db, user, h.location_id)
    db.delete(h); db.commit()
    return {"ok": True}


# ---------------------------------------------------------------------------
# Attendance
# ---------------------------------------------------------------------------
@app.get("/api/attendance/upload-template")
def upload_template(location_id: int, year: int = 0, month: int = 0,
                    fmt: str = "csv",
                    user: models.User = Depends(auth.get_current_user),
                    db: Session = Depends(get_db)):
    """Download a blank wide-format template (CSV or XLSX) pre-filled with
    this location's existing employees so the vendor can fill in days 1-31."""
    from calendar import monthrange
    from .engine.calculator import _merge, DEFAULT_RULE_CONFIG

    loc = _owned_location(db, user, location_id)
    cfg = _merge(DEFAULT_RULE_CONFIG, loc.rule_config)
    code_map = cfg.get("vendor_code_map", {})

    if not year or not month:
        today = date.today()
        year, month = today.year, today.month
    dm = monthrange(year, month)[1]

    info_headers = ["Month", "Vendor", "Emp Code", "Employee Name", "Designation",
                    "Gender", "Date of Joining", "Current Status", "Date of Exit"]
    day_headers = [str(d) for d in range(1, dm + 1)]
    headers = info_headers + day_headers

    month_label = date(year, month, 1).strftime("%b-%y")
    employees = db.query(models.Employee).filter_by(location_id=loc.id).all()

    rows = []
    for e in employees:
        status = "Exited" if e.exit_date else "Active"
        rows.append([
            month_label, loc.code, e.emp_code, e.name, e.title or "", e.gender,
            e.joining_date.strftime("%d-%b-%y") if e.joining_date else "",
            status,
            e.exit_date.strftime("%d-%b-%y") if e.exit_date else "",
        ] + [""] * dm)

    legend = "Vendor codes: " + ", ".join(f"{k}={v}" for k, v in code_map.items()) + \
             " (internal: A=Present, B=Half Present, C=Week Off, D=Holiday, E=Leave, F=Half Leave, G=Absent)"

    if fmt == "xlsx":
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        wb = Workbook(); ws = wb.active; ws.title = "Attendance Upload"
        ws.append(headers)
        hdr_fill = PatternFill("solid", fgColor="D9E1F2")
        for ci in range(1, len(headers) + 1):
            c = ws.cell(1, ci); c.fill = hdr_fill; c.font = Font(bold=True, size=9)
            c.alignment = Alignment(horizontal="center")
        for r in rows: ws.append(r)
        ws.append([])
        ws.append([legend])
        for ci in range(1, len(info_headers) + 1):
            ws.column_dimensions[ws.cell(1, ci).column_letter].width = 14
        for ci in range(len(info_headers) + 1, len(headers) + 1):
            ws.column_dimensions[ws.cell(1, ci).column_letter].width = 4.5
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        fname = f"attendance_template_{loc.code}_{year}_{month:02d}.xlsx"
        return StreamingResponse(buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename={fname}'})

    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(headers)
    for r in rows: w.writerow(r)
    w.writerow([])
    w.writerow([legend])
    fname = f"attendance_template_{loc.code}_{year}_{month:02d}.csv"
    return StreamingResponse(iter([out.getvalue()]), media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename={fname}'})


@app.post("/api/attendance/upload")
async def upload_attendance(
    location_id: int = Form(...),
    file: UploadFile = File(...),
    user: models.User = Depends(auth.get_current_user),
    db: Session = Depends(get_db),
):
    _owned_location(db, user, location_id)
    content = (await file.read()).decode("utf-8")
    reader = csv.DictReader(io.StringIO(content))
    emps = {e.emp_code: e for e in db.query(models.Employee).filter_by(location_id=location_id).all()}
    n = 0
    for row in reader:
        emp = emps.get(row["emp_code"])
        if not emp:
            continue
        d = datetime.strptime(row["day"], "%Y-%m-%d").date()
        db.query(models.AttendanceRecord).filter_by(employee_id=emp.id, day=d).delete()
        db.add(models.AttendanceRecord(
            employee_id=emp.id, day=d, code=row["code"].strip().upper(),
            worked_on_holiday=row.get("worked_on_holiday", "0") in ("1", "true", "True"),
        ))
        n += 1
    db.commit()
    return {"rows_ingested": n}


@app.post("/api/attendance/upload-wide")
async def upload_attendance_wide(
    location_id: int = Form(...),
    file: UploadFile = File(...),
    user: models.User = Depends(auth.get_current_user),
    db: Session = Depends(get_db),
):
    """Upload vendor wide-format sheet (one row per employee, days 1-31 as columns).
    Auto-creates employees not found. Maps vendor codes via location's vendor_code_map."""
    from calendar import monthrange
    from .engine.calculator import _merge, DEFAULT_RULE_CONFIG

    loc = _owned_location(db, user, location_id)
    cfg = _merge(DEFAULT_RULE_CONFIG, loc.rule_config)
    code_map = cfg.get("vendor_code_map", {})

    fname = (file.filename or "").lower()
    raw = await file.read()

    if fname.endswith((".xlsx", ".xls")):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
        rows_iter = list(ws.iter_rows(values_only=True))
        if not rows_iter:
            raise HTTPException(400, "Empty workbook")
        headers = [str(h or "").strip() for h in rows_iter[0]]
        data_rows = rows_iter[1:]
    else:
        content = raw.decode("utf-8")
        reader = list(csv.reader(io.StringIO(content)))
        if not reader:
            raise HTTPException(400, "Empty CSV")
        headers = [h.strip() for h in reader[0]]
        data_rows = reader[1:]

    hdr_lower = [h.lower() for h in headers]

    def _col(names):
        for n in names:
            if n.lower() in hdr_lower:
                return hdr_lower.index(n.lower())
        return None

    col_name = _col(["employee name", "name", "emp_name", "employee"])
    col_desig = _col(["designation", "title", "desig"])
    col_vendor = _col(["vendor", "vendor code", "vendor_code"])
    col_joining = _col(["date of joining", "joining", "joining_date", "doj"])
    col_status = _col(["current status", "status"])
    col_exit = _col(["date of exit", "dateofexit", "exit_date", "exit"])
    col_month = _col(["month"])
    col_gender = _col(["gender", "sex"])
    col_emp_code = _col(["emp code", "emp_code", "employee code", "empcode", "employee id"])

    day_cols = {}
    for i, h in enumerate(headers):
        try:
            d = int(h)
            if 1 <= d <= 31:
                day_cols[d] = i
        except ValueError:
            pass

    if col_name is None:
        raise HTTPException(400, "Missing 'Employee Name' column")
    if not day_cols:
        raise HTTPException(400, "No day columns (1-31) found in headers")

    emps = {e.name.strip().lower(): e for e in
            db.query(models.Employee).filter_by(location_id=loc.id).all()}
    if col_emp_code is not None:
        emps_by_code = {e.emp_code.strip().lower(): e for e in
                        db.query(models.Employee).filter_by(location_id=loc.id).all()}
    else:
        emps_by_code = {}

    def _parse_date(val):
        if not val:
            return None
        s = str(val).strip()
        for fmt in ("%d-%b-%y", "%d-%b-%Y", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                pass
        return None

    def _parse_month(val):
        if not val:
            return None, None
        s = str(val).strip()
        for fmt in ("%b-%y", "%b-%Y", "%B-%y", "%B-%Y", "%Y-%m", "%m-%Y"):
            try:
                dt = datetime.strptime(s, fmt)
                return dt.year, dt.month
            except ValueError:
                pass
        return None, None

    created = 0
    att_saved = 0
    skipped = []

    for row_num, row in enumerate(data_rows, start=2):
        row = list(row)
        if len(row) <= max(day_cols.values()):
            row += [None] * (max(day_cols.values()) + 1 - len(row))

        name_raw = str(row[col_name] or "").strip()
        if not name_raw:
            continue

        year, month = None, None
        if col_month is not None:
            year, month = _parse_month(row[col_month])
        if not year:
            raise HTTPException(400, f"Row {row_num}: cannot parse Month '{row[col_month] if col_month is not None else '(missing)'}'. Expected format: Apr-25")

        dm = monthrange(year, month)[1]

        emp_code_raw = str(row[col_emp_code]).strip() if col_emp_code is not None and row[col_emp_code] else None
        emp = None
        if emp_code_raw:
            emp = emps_by_code.get(emp_code_raw.lower())
        if not emp:
            emp = emps.get(name_raw.lower())

        if not emp:
            joining = _parse_date(row[col_joining]) if col_joining is not None else None
            exit_d = _parse_date(row[col_exit]) if col_exit is not None else None
            title = str(row[col_desig] or "").strip() if col_desig is not None else None
            gender_raw = str(row[col_gender] or "").strip().upper() if col_gender is not None else "M"
            gender = gender_raw[0] if gender_raw and gender_raw[0] in ("M", "F", "O") else "M"

            auto_code = emp_code_raw or f"AUTO-{row_num}"
            emp = models.Employee(
                location_id=loc.id,
                emp_code=auto_code,
                title=title or None,
                name=name_raw,
                gender=gender,
                joining_date=joining or date(year, month, 1),
                exit_date=exit_d,
                monthly_salary=0,
                opening_leave=0,
            )
            db.add(emp); db.flush()
            emps[name_raw.lower()] = emp
            if emp_code_raw:
                emps_by_code[emp_code_raw.lower()] = emp
            created += 1
        else:
            if col_exit is not None:
                exit_d = _parse_date(row[col_exit])
                status_raw = str(row[col_status] or "").strip().lower() if col_status is not None else ""
                if status_raw in ("exited", "inactive", "resigned", "terminated") or exit_d:
                    emp.exit_date = exit_d

        for day_num, col_idx in day_cols.items():
            if day_num > dm:
                continue
            vendor_code = str(row[col_idx] or "").strip().upper()
            if not vendor_code:
                continue
            internal = code_map.get(vendor_code, code_map.get(vendor_code.lower(), ""))
            if not internal:
                internal = vendor_code if vendor_code in "ABCDEFG" else ""
            if not internal or internal not in "ABCDEFG":
                skipped.append(f"Row {row_num} day {day_num}: unknown code '{vendor_code}'")
                continue
            dt = date(year, month, day_num)
            db.query(models.AttendanceRecord).filter_by(employee_id=emp.id, day=dt).delete()
            db.add(models.AttendanceRecord(
                employee_id=emp.id, day=dt, code=internal,
                worked_on_holiday=False,
            ))
            att_saved += 1

    db.commit()
    return {
        "rows_processed": len(data_rows),
        "employees_created": created,
        "attendance_saved": att_saved,
        "skipped": skipped[:20],
    }


@app.get("/api/attendance")
def get_attendance(location_id: int, year: int, month: int,
                   user: models.User = Depends(auth.get_current_user),
                   db: Session = Depends(get_db)):
    """Grid data for the attendance entry UI."""
    from calendar import monthrange
    from .engine.calculator import _is_scheduled_week_off, _merge, DEFAULT_RULE_CONFIG

    loc = _owned_location(db, user, location_id)
    cfg = _merge(DEFAULT_RULE_CONFIG, loc.rule_config)
    dm = monthrange(year, month)[1]
    first, last = date(year, month, 1), date(year, month, dm)
    emps = db.query(models.Employee).filter_by(location_id=loc.id).all()
    hols = {h.day: h.name for h in db.query(models.Holiday)
            .filter(models.Holiday.location_id == loc.id,
                    models.Holiday.day >= first, models.Holiday.day <= last).all()}

    recs = {(r.employee_id, r.day): r for r in db.query(models.AttendanceRecord)
            .filter(models.AttendanceRecord.day >= first, models.AttendanceRecord.day <= last,
                    models.AttendanceRecord.employee_id.in_([e.id for e in emps])).all()}

    grid = []
    for e in emps:
        row = {"employee_id": e.id, "emp_code": e.emp_code, "name": e.name,
               "joining_date": e.joining_date.isoformat(),
               "exit_date": e.exit_date.isoformat() if e.exit_date else None,
               "days": []}
        for d in range(1, dm + 1):
            dt = date(year, month, d)
            r = recs.get((e.id, dt))
            if r:
                code = r.code
                worked = r.worked_on_holiday
            else:
                if dt in hols:
                    code = "D"
                elif _is_scheduled_week_off(dt, cfg, e.week_off_day):
                    code = "C"
                elif (e.joining_date and dt < e.joining_date) or (e.exit_date and dt > e.exit_date):
                    code = ""  # out of employment window
                else:
                    code = "A"
                worked = False
            row["days"].append({"day": dt.isoformat(), "code": code,
                                "worked_on_holiday": worked,
                                "holiday_name": hols.get(dt)})
        grid.append(row)
    return {"year": year, "month": month, "dm": dm,
            "holidays": {k.isoformat(): v for k, v in hols.items()},
            "rows": grid}


@app.get("/api/attendance/tracker")
def attendance_tracker(location_id: int, year: int, month: int,
                       user: models.User = Depends(auth.get_current_user),
                       db: Session = Depends(get_db)):
    """Simple tracker view: each cell is a single letter — P/A/W/H/L/—.

    P = Present (A or B)
    A = Absent (G)
    W = Week-off (C)
    H = Holiday (D)
    L = Leave (E or F)
    —  = out of employment window
    """
    from calendar import monthrange
    loc = _owned_location(db, user, location_id)
    dm = monthrange(year, month)[1]
    first, last = date(year, month, 1), date(year, month, dm)
    emps = db.query(models.Employee).filter_by(location_id=loc.id).all()
    hols = {h.day: h.name for h in db.query(models.Holiday)
            .filter(models.Holiday.location_id == loc.id,
                    models.Holiday.day >= first, models.Holiday.day <= last).all()}
    recs = {(r.employee_id, r.day): r for r in db.query(models.AttendanceRecord)
            .filter(models.AttendanceRecord.day >= first,
                    models.AttendanceRecord.day <= last,
                    models.AttendanceRecord.employee_id.in_([e.id for e in emps])).all()}

    def _tracker_letter(code: str) -> str:
        return {"A": "P", "B": "P", "C": "W", "D": "H",
                "E": "L", "F": "L", "G": "A"}.get(code, "")

    rows = []
    for e in emps:
        summary = {"P": 0.0, "A": 0.0, "W": 0.0, "H": 0.0, "L": 0.0}
        days = []
        for d in range(1, dm + 1):
            dt = date(year, month, d)
            if (e.joining_date and dt < e.joining_date) or (e.exit_date and dt > e.exit_date):
                days.append("—")
                continue
            r = recs.get((e.id, dt))
            if not r:
                days.append("")
                continue
            letter = _tracker_letter(r.code)
            days.append(letter)
            if letter == "P" and r.code == "B":
                summary["P"] += 0.5
            elif letter == "L" and r.code == "F":
                summary["L"] += 0.5
            elif letter in summary:
                summary[letter] += 1
        rows.append({
            "employee_id": e.id, "emp_code": e.emp_code,
            "title": e.title or "", "name": e.name, "gender": e.gender,
            "days": days, "summary": summary,
        })
    return {"year": year, "month": month, "dm": dm,
            "holidays": {k.isoformat(): v for k, v in hols.items()},
            "rows": rows}


@app.post("/api/attendance/save")
async def save_attendance(request: Request,
                          user: models.User = Depends(auth.get_current_user),
                          db: Session = Depends(get_db)):
    """Save one grid page — body: {location_id, rows: [{employee_id, days:[{day,code,worked_on_holiday}]}]}"""
    body = await request.json()
    _owned_location(db, user, body["location_id"])
    allowed = {e.id for e in db.query(models.Employee)
               .filter_by(location_id=body["location_id"]).all()}
    n = 0
    for row in body["rows"]:
        eid = int(row["employee_id"])
        if eid not in allowed:
            continue
        for d in row["days"]:
            day = datetime.strptime(d["day"], "%Y-%m-%d").date()
            code = (d.get("code") or "").strip().upper()
            db.query(models.AttendanceRecord).filter_by(employee_id=eid, day=day).delete()
            if code:
                db.add(models.AttendanceRecord(
                    employee_id=eid, day=day, code=code,
                    worked_on_holiday=bool(d.get("worked_on_holiday", False)),
                ))
                n += 1
    db.commit()
    return {"saved": n}


# ---------------------------------------------------------------------------
# Payroll
# ---------------------------------------------------------------------------
@app.post("/api/payroll/run")
def run_payroll(body: schemas.PayrollRunIn,
                user: models.User = Depends(auth.get_current_user),
                db: Session = Depends(get_db)):
    loc = _owned_location(db, user, body.location_id)
    run = models.PayrollRun(location_id=loc.id, month=body.month, year=body.year)
    db.add(run); db.commit(); db.refresh(run)

    employees = db.query(models.Employee).filter_by(location_id=loc.id).all()
    first = date(body.year, body.month, 1)
    last = date(body.year + (1 if body.month == 12 else 0),
                1 if body.month == 12 else body.month + 1, 1)

    # Merge declared holidays into attendance stream as D if not already recorded
    holidays = {h.day for h in db.query(models.Holiday).filter_by(location_id=loc.id).all()}

    results = []
    for e in employees:
        recs = db.query(models.AttendanceRecord).filter(
            models.AttendanceRecord.employee_id == e.id,
            models.AttendanceRecord.day >= first,
            models.AttendanceRecord.day < last,
        ).all()
        rec_dicts = [{"day": r.day, "code": r.code, "worked_on_holiday": r.worked_on_holiday}
                     for r in recs]
        seen = {r["day"] for r in rec_dicts}
        for h in holidays:
            if first <= h < last and h not in seen:
                rec_dicts.append({"day": h, "code": "D", "worked_on_holiday": False})

        payload = calculate_payroll(
            employee={
                "emp_code": e.emp_code, "name": e.name, "gender": e.gender,
                "title": e.title,
                "joining_date": e.joining_date, "exit_date": e.exit_date,
                "monthly_salary": e.monthly_salary, "opening_leave": e.opening_leave,
                "week_off_day": e.week_off_day,
            },
            year=body.year, month=body.month, records=rec_dicts,
            rule_config=loc.rule_config,
        )
        db.add(models.PayrollResult(run_id=run.id, employee_id=e.id, payload=payload))
        results.append(payload)
    db.commit()
    return {"run_id": run.id, "results": results}


@app.get("/api/payroll/runs")
def list_runs(location_id: int | None = None,
              user: models.User = Depends(auth.get_current_user),
              db: Session = Depends(get_db)):
    loc_ids = [l.id for l in db.query(models.Location).filter_by(company_id=user.company_id).all()]
    q = db.query(models.PayrollRun).filter(models.PayrollRun.location_id.in_(loc_ids))
    if location_id:
        _owned_location(db, user, location_id)
        q = q.filter(models.PayrollRun.location_id == location_id)
    return [{"id": r.id, "location_id": r.location_id, "month": r.month,
             "year": r.year, "created_at": r.created_at.isoformat()} for r in q.all()]


@app.get("/api/payroll/report")
def payroll_report(run_id: int, format: str = "csv",
                   user: models.User = Depends(auth.get_current_user),
                   db: Session = Depends(get_db)):
    run = db.get(models.PayrollRun, run_id)
    if not run:
        raise HTTPException(404)
    _owned_location(db, user, run.location_id)
    rows = [r.payload for r in run.results]
    headers = [
        "Emp Code", "Title", "Name", "Gender",
        "Total Days", "Payable Days",
        "Present", "Half Present", "Week Off", "Holiday",
        "Leave", "Half Leave", "Absent", "Holiday Work",
        "Approved Leave", "Leave Bucket", "Closing Leave",
        "Monthly Salary", "Per Day", "Gross Payable",
        "Flags", "Notes",
    ]

    def _row_values(r):
        return [
            r["employee"]["emp_code"], r["employee"].get("title") or "",
            r["employee"]["name"], r["employee"]["gender"],
            r["H_total_days"], r["I_payable_days"],
            r["counts"]["A"], r["counts"]["B"], r["counts"]["C"], r["counts"]["D"],
            r["counts"]["E"], r["counts"]["F"], r["counts"]["G"], r["counts"]["J"],
            r["approved_leave"], r["leave_bucket"], r["leave_ledger"]["L_closing"],
            r["salary"]["monthly"], r["salary"]["per_day"], r["salary"]["gross_payable"],
            "; ".join(r["flags"]), "; ".join(r["notes"]),
        ]

    if format == "xlsx":
        from openpyxl import Workbook
        wb = Workbook(); ws = wb.active; ws.title = "Payroll"
        ws.append(headers)
        for r in rows: ws.append(_row_values(r))
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return StreamingResponse(buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename=payroll_{run_id}.xlsx'})

    out = io.StringIO()
    w = csv.writer(out); w.writerow(headers)
    for r in rows: w.writerow(_row_values(r))
    return StreamingResponse(iter([out.getvalue()]), media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename=payroll_{run_id}.csv'})


@app.get("/api/payroll/checked-report")
def checked_report(run_id: int,
                   user: models.User = Depends(auth.get_current_user),
                   db: Session = Depends(get_db)):
    """Wide-format checked report: employee info + days 1-31 as vendor codes + summary columns."""
    from calendar import monthrange
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from .engine.calculator import _merge, DEFAULT_RULE_CONFIG

    run = db.get(models.PayrollRun, run_id)
    if not run:
        raise HTTPException(404)
    loc = _owned_location(db, user, run.location_id)
    cfg = _merge(DEFAULT_RULE_CONFIG, loc.rule_config)
    vendor_map = cfg.get("vendor_code_map", {})
    reverse_map = {v: k for k, v in vendor_map.items()}

    dm = monthrange(run.year, run.month)[1]
    first = date(run.year, run.month, 1)
    last = date(run.year, run.month, dm)

    payroll_results = {pr.employee_id: pr.payload for pr in run.results}
    employees = {e.id: e for e in
                 db.query(models.Employee).filter_by(location_id=loc.id).all()}

    recs_all = db.query(models.AttendanceRecord).filter(
        models.AttendanceRecord.day >= first,
        models.AttendanceRecord.day <= last,
        models.AttendanceRecord.employee_id.in_(list(employees.keys())),
    ).all()
    recs_by_emp = {}
    for r in recs_all:
        recs_by_emp.setdefault(r.employee_id, {})[r.day] = r

    wb = Workbook()
    ws = wb.active
    ws.title = "Checked Report"

    info_headers = ["Emp Code", "Vendor", "Employee Name", "Designation",
                    "Date of Joining", "Current Status", "Date of Exit"]
    day_headers = [str(d) for d in range(1, dm + 1)]
    summary_headers = [
        "Present (Full Day)", "Present (Half Day)", "Week Off",
        "Declared Holiday", "Leave (Full Day)", "Leave (Half Day)",
        "Absent", "Total Days", "Payable Days",
        "Leave With Pay", "Leave Without Pay", "Total Payable Days",
    ]
    all_headers = info_headers + day_headers + summary_headers

    hdr_fill = PatternFill("solid", fgColor="D9E1F2")
    hdr_font = Font(bold=True, size=9)
    thin = Side(style="thin", color="B0B0B0")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ci, h in enumerate(all_headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = cell_border

    code_fills = {
        "P": PatternFill("solid", fgColor="C6EFCE"),
        "A": PatternFill("solid", fgColor="FFC7CE"),
        "WO": PatternFill("solid", fgColor="BDD7EE"),
        "H": PatternFill("solid", fgColor="FCE4D6"),
        "L": PatternFill("solid", fgColor="FFF2CC"),
        "HL": PatternFill("solid", fgColor="FFF2CC"),
        "HD": PatternFill("solid", fgColor="C6EFCE"),
    }

    row_idx = 2
    for eid, payload in payroll_results.items():
        emp = employees.get(eid)
        if not emp:
            continue

        status = "Active" if not emp.exit_date else "Exited"
        info = [
            emp.emp_code,
            loc.code,
            emp.name,
            emp.title or "",
            emp.joining_date.strftime("%d-%b-%y") if emp.joining_date else "",
            status,
            emp.exit_date.strftime("%d-%b-%y") if emp.exit_date else "",
        ]

        day_recs = recs_by_emp.get(eid, {})
        day_codes = []
        for d in range(1, dm + 1):
            dt = date(run.year, run.month, d)
            rec = day_recs.get(dt)
            if rec:
                vendor_code = reverse_map.get(rec.code, rec.code)
            else:
                if (emp.joining_date and dt < emp.joining_date) or (emp.exit_date and dt > emp.exit_date):
                    vendor_code = ""
                else:
                    vendor_code = ""
            day_codes.append(vendor_code)

        counts = payload.get("counts", {})
        leave_taken = counts.get("E", 0) + counts.get("F", 0)
        approved_leave = payload.get("approved_leave", 0)
        leave_bucket = payload.get("leave_bucket", 0)
        leave_with_pay = min(leave_taken, leave_bucket)
        leave_without_pay = max(0, leave_taken - leave_bucket)

        payable_days = (counts.get("A", 0) + counts.get("B", 0) +
                        counts.get("C", 0) + counts.get("D", 0) +
                        counts.get("E", 0) + counts.get("F", 0))
        total_payable = payable_days - leave_without_pay

        summary = [
            counts.get("A", 0),
            counts.get("B", 0),
            counts.get("C", 0),
            counts.get("D", 0),
            counts.get("E", 0),
            counts.get("F", 0),
            counts.get("G", 0),
            payload.get("H_total_days", 0),
            payable_days,
            round(leave_with_pay, 1),
            round(leave_without_pay, 1),
            round(total_payable, 1),
        ]

        row_data = info + day_codes + summary
        for ci, val in enumerate(row_data, 1):
            c = ws.cell(row=row_idx, column=ci, value=val)
            c.border = cell_border
            c.alignment = Alignment(horizontal="center")
            c.font = Font(size=9)
            if len(info_headers) < ci <= len(info_headers) + dm:
                vc = str(val).strip().upper()
                if vc in code_fills:
                    c.fill = code_fills[vc]
                    c.font = Font(size=9, bold=True)

        for ci in range(1, len(info) + 1):
            ws.cell(row=row_idx, column=ci).alignment = Alignment(horizontal="left")

        row_idx += 1

    for ci in range(1, len(info_headers) + 1):
        ws.column_dimensions[ws.cell(1, ci).column_letter].width = 14
    for ci in range(len(info_headers) + 1, len(info_headers) + dm + 1):
        ws.column_dimensions[ws.cell(1, ci).column_letter].width = 4.5
    for ci in range(len(info_headers) + dm + 1, len(all_headers) + 1):
        ws.column_dimensions[ws.cell(1, ci).column_letter].width = 12

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"checked_report_{loc.code}_{run.year}_{run.month:02d}.xlsx"
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename={fname}'})
